"""
BandiRadar v2 — Aggiornamento automatico settimanale
Zero costi: scraping fonti ufficiali + Excel + GitHub Pages

Eseguito ogni lunedi da GitHub Actions (gratuito).
"""

import json
import os
import sys
import re
from datetime import date, datetime
from pathlib import Path

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Percorsi ──────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
DATA_FILE   = BASE_DIR / "data" / "bandi_base.json"
OUTPUT_XLSX = BASE_DIR / "output" / "BandiRadar.xlsx"
OUTPUT_LOG  = BASE_DIR / "output" / "ultimo_aggiornamento.txt"
OUTPUT_JSON = BASE_DIR / "output" / "bandi.json"   # per la web page

TODAY = date.today()

# ── Fonti da scrapare (HTML semplice, no JS rendering) ────
SOURCES = [
    {
        "nome": "CCIAA Toscana Nord-Ovest",
        "url": "https://tno.camcom.it/bandi-alle-imprese",
        "regione": "Regione Toscana",
        "settore_default": "Digitalizzazione / Transizione 4.0",
    },
    {
        "nome": "Regione Lombardia Bandi",
        "url": "https://www.bandi.regione.lombardia.it/servizi/servizio/bandi?categoria=1",
        "regione": "Regione Lombardia",
        "settore_default": "Innovazione / R&S",
    },
    {
        "nome": "incentivi.gov.it",
        "url": "https://www.incentivi.gov.it/it/catalogo?pmi=true&stato=aperto",
        "regione": "Nazionale",
        "settore_default": "Digitalizzazione / Transizione 4.0",
    },
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; BandiRadarBot/2.0; +https://github.com/bandiradar)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "it-IT,it;q=0.9",
}

KEYWORDS = [
    "digitalizzazione", "transizione", "innovazione", "cybersecurity",
    "voucher", "formazione", "export", "e-commerce", "simest",
    "digitale", "4.0", "pmi", "imprese", "contributo", "fondo perduto",
    "branding", "ui", "ux", "consulenza",
]

SETTORI_MAP = {
    "digitaliz": "Digitalizzazione / Transizione 4.0",
    "transizione": "Digitalizzazione / Transizione 4.0",
    "4.0": "Digitalizzazione / Transizione 4.0",
    "5.0": "Digitalizzazione / Transizione 4.0",
    "cyber": "Cybersecurity",
    "cloud": "Cybersecurity",
    "sicurezza": "Cybersecurity",
    "innovazione": "Innovazione / R&S",
    "ricerca": "Innovazione / R&S",
    "brevett": "Innovazione / R&S",
    "export": "E-commerce & Export",
    "e-commerce": "E-commerce & Export",
    "internazional": "E-commerce & Export",
    "fiere": "E-commerce & Export",
    "formazione": "Formazione 4.0",
    "competenze": "Formazione 4.0",
    "consulenza": "Voucher Consulenza / Brand / UI-UX",
    "voucher": "Voucher Consulenza / Brand / UI-UX",
    "brand": "Voucher Consulenza / Brand / UI-UX",
    "ui/ux": "Voucher Consulenza / Brand / UI-UX",
    "marchi": "Voucher Consulenza / Brand / UI-UX",
}


# ── Calcolo status ────────────────────────────────────────
def compute_status(scadenza_str):
    try:
        sc = date.fromisoformat(str(scadenza_str))
        dl = (sc - TODAY).days
        if dl < 0:    return "CHIUSO"
        if dl <= 30:  return "IN SCADENZA"
        return "APERTO"
    except Exception:
        return "APERTO"

def days_left(scadenza_str):
    try:
        return (date.fromisoformat(str(scadenza_str)) - TODAY).days
    except Exception:
        return 999

def guess_settore(testo):
    testo_lower = testo.lower()
    for keyword, settore in SETTORI_MAP.items():
        if keyword in testo_lower:
            return settore
    return "Digitalizzazione / Transizione 4.0"

def extract_date(testo):
    patterns = [
        r"\b(\d{1,2})[/\-](\d{1,2})[/\-](20\d{2})\b",
        r"\b(20\d{2})[/\-](\d{1,2})[/\-](\d{1,2})\b",
    ]
    for pat in patterns:
        m = re.search(pat, testo)
        if m:
            try:
                g = m.groups()
                if len(g[0]) == 4:
                    d = date(int(g[0]), int(g[1]), int(g[2]))
                else:
                    d = date(int(g[2]), int(g[1]), int(g[0]))
                if d.year >= TODAY.year:
                    return d.isoformat()
            except Exception:
                pass
    return None


# ── Scraping ──────────────────────────────────────────────
def scrape_source(source):
    """Scarica e analizza una fonte, restituisce lista di bandi trovati."""
    trovati = []
    try:
        r = httpx.get(source["url"], headers=HEADERS, timeout=15, follow_redirects=True)
        if r.status_code != 200:
            print(f"  [{source['nome']}] HTTP {r.status_code} — skip")
            return []

        soup = BeautifulSoup(r.text, "html.parser")

        # Cerca link a bandi/contributi nella pagina
        for a in soup.find_all("a", href=True):
            testo = a.get_text(strip=True)
            href = a["href"]

            if len(testo) < 10 or len(testo) > 200:
                continue

            # Filtra per keyword rilevanti
            testo_lower = testo.lower()
            if not any(k in testo_lower for k in KEYWORDS):
                continue

            # Salta link di navigazione generici
            if any(skip in testo_lower for skip in ["cookie", "privacy", "accessibilità", "home", "contatti", "login"]):
                continue

            # Costruisci URL assoluto
            if href.startswith("http"):
                url_bando = href
            elif href.startswith("/"):
                from urllib.parse import urlparse
                base = urlparse(source["url"])
                url_bando = f"{base.scheme}://{base.netloc}{href}"
            else:
                continue

            # Cerca data di scadenza nel testo circostante
            parent_text = ""
            parent = a.parent
            for _ in range(3):
                if parent:
                    parent_text += parent.get_text(" ", strip=True)
                    parent = parent.parent if parent else None

            scadenza = extract_date(parent_text) or f"{TODAY.year}-12-31"

            # Filtra anni non plausibili
            if scadenza and int(scadenza[:4]) < TODAY.year:
                scadenza = f"{TODAY.year}-12-31"

            trovati.append({
                "titolo": testo[:120],
                "ente": source["nome"],
                "regione": source["regione"],
                "settore": guess_settore(testo + " " + parent_text),
                "contributo": "Vedere bando",
                "intensita": "Vedere bando",
                "scadenza": scadenza,
                "descrizione": testo[:150],
                "link": url_bando,
                "note": f"Trovato automaticamente su {source['nome']} — verificare prima dell'uso",
                "fonte": "AUTO",
            })

        print(f"  [{source['nome']}] {len(trovati)} elementi trovati")
        return trovati[:10]  # max 10 per fonte

    except Exception as e:
        print(f"  [{source['nome']}] Errore: {e}")
        return []


# ── Merge bandi ───────────────────────────────────────────
def merge_bandi(base, scraped):
    """Unisce bandi verificati + bandi trovati via scraping, senza duplicati."""
    titoli_noti = {b["titolo"].lower().strip() for b in base}
    nuovi = []
    for b in scraped:
        titolo_norm = b["titolo"].lower().strip()
        # Evita duplicati per titolo simile
        if not any(titolo_norm in noto or noto in titolo_norm for noto in titoli_noti):
            nuovi.append(b)
            titoli_noti.add(titolo_norm)

    # Aggiorna status da date reali per tutti
    tutti = base + nuovi
    for b in tutti:
        b["status"] = compute_status(b.get("scadenza", ""))

    # Rimuovi bandi scaduti da più di 60 giorni
    tutti = [b for b in tutti if days_left(b.get("scadenza", "")) > -60]

    # Ordina: verificati prima, poi per scadenza
    tutti.sort(key=lambda b: (
        0 if b.get("fonte") == "VERIFICATO" else 1,
        b.get("scadenza", "9999")
    ))

    return tutti


# ── Excel ─────────────────────────────────────────────────
def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
def border():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

C = {
    "navy":       "0F1C35",
    "blue_hdr":   "1E3A5F",
    "white":      "FFFFFF",
    "light":      "F8FAFC",
    "green_bg":   "D1FAE5",
    "green_tx":   "065F46",
    "yellow_bg":  "FEF3C7",
    "yellow_tx":  "92400E",
    "red_bg":     "FEE2E2",
    "red_tx":     "991B1B",
    "purple_bg":  "EDE9FE",
    "purple_tx":  "5B21B6",
}

def status_colors(b):
    s = b.get("status", "APERTO")
    dl = days_left(b.get("scadenza", ""))
    if s == "CHIUSO":     return C["red_bg"],    C["red_tx"],    "CHIUSO"
    if s == "IN SCADENZA" or (dl >= 0 and dl <= 30):
        return C["yellow_bg"], C["yellow_tx"], "IN SCADENZA"
    return C["green_bg"],  C["green_tx"],  "APERTO"

def genera_excel(bandi, filepath):
    today_str = TODAY.strftime("%d/%m/%Y")
    wb = Workbook()

    # ── FOGLIO 1: BANDI ATTIVI ────────────────────────────
    ws = wb.active
    ws.title = "Bandi Attivi"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Intestazione
    for row, (testo, ht, bold, sz) in enumerate([
        (f"BANDIRADAR — Monitor Bandi & Contributi per PMI | Aggiornato: {today_str}", 30, True, 14),
        (f"Dataset: {sum(1 for b in bandi if b.get('fonte')=='VERIFICATO')} bandi verificati + "
         f"{sum(1 for b in bandi if b.get('fonte')!='VERIFICATO')} trovati automaticamente | "
         f"Link pubblico: github.com/TUO-USERNAME/bandi-radar", 20, False, 9),
    ], 1):
        ws.merge_cells(f"A{row}:L{row}")
        c = ws.cell(row=row, column=1, value=testo)
        c.font = Font(name="Arial", bold=bold, size=sz, color=C["white"])
        c.fill = fill(C["navy"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = ht

    # Header colonne
    COLS = [
        ("TITOLO BANDO",     38),
        ("ENTE",             26),
        ("REGIONE",          18),
        ("SETTORE",          26),
        ("CONTRIBUTO",       20),
        ("INTENSITA",        22),
        ("SCADENZA",         13),
        ("GG",               8),
        ("STATO",            14),
        ("FONTE",            10),
        ("NOTE",             35),
        ("LINK",             12),
    ]
    for i, (h, w) in enumerate(COLS, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=C["white"])
        c.fill = fill(C["blue_hdr"])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 26

    # Dati
    for ri, b in enumerate(bandi, 4):
        bg_c, tx_c, lbl = status_colors(b)
        dl = days_left(b.get("scadenza", ""))
        try:
            sc_f = date.fromisoformat(b["scadenza"]).strftime("%d/%m/%Y")
        except Exception:
            sc_f = b.get("scadenza", "")
        dl_str = str(dl) + "gg" if dl >= 0 else "SCAD."
        is_verificato = b.get("fonte") == "VERIFICATO"

        row_bg = fill(C["white"]) if ri % 2 == 0 else fill(C["light"])
        vals = [
            b.get("titolo", ""),
            b.get("ente", ""),
            b.get("regione", ""),
            b.get("settore", ""),
            b.get("contributo", ""),
            b.get("intensita", ""),
            sc_f,
            dl_str,
            lbl,
            "OK" if is_verificato else "AUTO",
            b.get("note", ""),
            b.get("link", ""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = Font(name="Arial", size=9)
            cell.fill = row_bg
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border()

        # Stato colorato
        sc = ws.cell(row=ri, column=9)
        sc.fill = fill(bg_c); sc.font = Font(name="Arial", bold=True, size=9, color=tx_c)
        sc.alignment = Alignment(horizontal="center", vertical="center")

        # GG rosso se urgente
        if 0 <= dl <= 30:
            dc = ws.cell(row=ri, column=8)
            dc.fill = fill(C["yellow_bg"]); dc.font = Font(name="Arial", bold=True, size=9, color=C["yellow_tx"])
            dc.alignment = Alignment(horizontal="center", vertical="center")

        # Fonte: verde=verificato, viola=auto
        fc = ws.cell(row=ri, column=10)
        if is_verificato:
            fc.fill = fill(C["green_bg"]); fc.font = Font(name="Arial", bold=True, size=9, color=C["green_tx"])
        else:
            fc.fill = fill(C["purple_bg"]); fc.font = Font(name="Arial", bold=True, size=9, color=C["purple_tx"])
        fc.alignment = Alignment(horizontal="center", vertical="center")

        # Link ipertestuale
        lc = ws.cell(row=ri, column=12)
        if b.get("link"):
            lc.hyperlink = b["link"]
            lc.font = Font(name="Arial", size=9, color="2563EB", underline="single")
            lc.value = "Apri"

        ws.row_dimensions[ri].height = 36

    ws.auto_filter.ref = f"A3:L{3+len(bandi)}"

    # Riga totale
    tr = 4 + len(bandi)
    ws.merge_cells(f"A{tr}:L{tr}")
    tc = ws[f"A{tr}"]
    ap = sum(1 for b in bandi if b.get("status") == "APERTO")
    ins = sum(1 for b in bandi if b.get("status") == "IN SCADENZA")
    verif = sum(1 for b in bandi if b.get("fonte") == "VERIFICATO")
    tc.value = f"  {len(bandi)} bandi | {ap} aperti | {ins} in scadenza | {verif} verificati manualmente | Aggiornato: {today_str}"
    tc.font = Font(name="Arial", bold=True, size=9, color=C["white"])
    tc.fill = fill(C["navy"])
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[tr].height = 20

    # ── FOGLIO 2: DASHBOARD ───────────────────────────────
    ws2 = wb.create_sheet("Dashboard")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:E1")
    c = ws2["A1"]
    c.value = f"DASHBOARD — {today_str}"
    c.font = Font(name="Arial", bold=True, size=14, color=C["white"])
    c.fill = fill(C["navy"]); c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    # Bandi urgenti (<=30 giorni)
    ws2.merge_cells("A3:E3")
    ws2["A3"].value = "BANDI IN SCADENZA ENTRO 30 GIORNI"
    ws2["A3"].font = Font(name="Arial", bold=True, size=12, color=C["yellow_tx"])
    ws2["A3"].fill = fill(C["yellow_bg"])
    ws2["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[3].height = 26

    urgenti = [(b, days_left(b.get("scadenza",""))) for b in bandi if 0 <= days_left(b.get("scadenza","")) <= 30]
    for i, h in enumerate(["BANDO", "ENTE", "SCADENZA", "GIORNI", "LINK"], 1):
        c = ws2.cell(row=4, column=i, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color=C["white"])
        c.fill = fill(C["yellow_tx"]); c.alignment = Alignment(horizontal="center"); c.border = border()
        ws2.column_dimensions[get_column_letter(i)].width = [38, 26, 13, 10, 40][i-1]

    if urgenti:
        for ri2, (b, dl) in enumerate(sorted(urgenti, key=lambda x: x[1]), 5):
            try: sc_f = date.fromisoformat(b["scadenza"]).strftime("%d/%m/%Y")
            except: sc_f = b.get("scadenza","")
            vals = [b.get("titolo",""), b.get("ente",""), sc_f, str(dl)+" gg", b.get("link","")]
            for ci, v in enumerate(vals, 1):
                cell = ws2.cell(row=ri2, column=ci, value=v)
                cell.fill = fill(C["yellow_bg"]); cell.font = Font(name="Arial", bold=True, size=10, color=C["yellow_tx"])
                cell.alignment = Alignment(vertical="center", wrap_text=True); cell.border = border()
            if b.get("link"):
                lc = ws2.cell(row=ri2, column=5)
                lc.hyperlink = b["link"]; lc.font = Font(name="Arial", size=9, color="2563EB", underline="single")
                lc.value = "Apri bando"
            ws2.row_dimensions[ri2].height = 28
    else:
        ws2.merge_cells("A5:E5")
        ws2["A5"].value = "Nessun bando in scadenza imminente questa settimana."
        ws2["A5"].font = Font(name="Arial", italic=True, size=10, color="94A3B8")

    # ── FOGLIO 3: LEGENDA ─────────────────────────────────
    ws3 = wb.create_sheet("Info & Fonti")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:C1")
    ws3["A1"].value = "LEGENDA E FONTI MONITORATE"
    ws3["A1"].font = Font(name="Arial", bold=True, size=13, color=C["white"])
    ws3["A1"].fill = fill(C["navy"]); ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["C"].width = 14

    legenda = [
        ("COLONNA FONTE", "", ""),
        ("OK (verde)", "Bando verificato manualmente su fonte ufficiale — affidabile", ""),
        ("AUTO (viola)", "Trovato automaticamente via scraping — verificare prima dell'uso con clienti", ""),
        ("", "", ""),
        ("COLONNA STATO", "", ""),
        ("APERTO (verde)", "Bando con scadenza futura > 30 giorni", ""),
        ("IN SCADENZA (giallo)", "Scadenza entro 30 giorni — agire subito", ""),
        ("CHIUSO (rosso)", "Bando scaduto (mostrato per 60 giorni poi rimosso)", ""),
        ("", "", ""),
        ("FONTI MONITORATE", "", "FREQ."),
        ("Regione Toscana", "regione.toscana.it / bandi FESR 2021-2027", "Sett."),
        ("Regione Lombardia", "bandi.regione.lombardia.it", "Sett."),
        ("MIMIT", "mimit.gov.it/it/incentivi", "Mens."),
        ("SIMEST", "simest.it (tutte le linee Fondo 394)", "Mens."),
        ("CCIAA Toscana NW", "tno.camcom.it/bandi (Lucca-Massa-Pisa)", "Mens."),
        ("CCIAA Arezzo-Siena", "as.camcom.it", "Mens."),
        ("CCIAA Firenze", "fi.camcom.gov.it", "Mens."),
        ("EUIPO", "euipo.europa.eu/sme-fund/2026", "Mens."),
        ("incentivi.gov.it", "Catalogo MIMIT (aggregatore ufficiale)", "Sett."),
        ("Agenzia Entrate", "Crediti d'imposta Formazione e Beni 4.0", "Mens."),
    ]
    for ri, (a, b_txt, c_txt) in enumerate(legenda, 3):
        ws3.cell(row=ri, column=1, value=a).font = Font(name="Arial", bold=bool(a and not b_txt), size=10)
        ws3.cell(row=ri, column=2, value=b_txt).font = Font(name="Arial", size=10)
        ws3.cell(row=ri, column=3, value=c_txt).font = Font(name="Arial", size=10)
        ws3.row_dimensions[ri].height = 20

    OUTPUT_XLSX.parent.mkdir(exist_ok=True)
    wb.save(filepath)
    print(f"Excel salvato: {filepath} ({len(bandi)} bandi)")


# ── Salva JSON per web page ───────────────────────────────
def salva_json(bandi):
    data = {
        "aggiornato": TODAY.isoformat(),
        "totale": len(bandi),
        "aperti": sum(1 for b in bandi if b.get("status") == "APERTO"),
        "in_scadenza": sum(1 for b in bandi if b.get("status") == "IN SCADENZA"),
        "verificati": sum(1 for b in bandi if b.get("fonte") == "VERIFICATO"),
        "bandi": bandi,
    }
    OUTPUT_JSON.parent.mkdir(exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"JSON salvato: {OUTPUT_JSON}")


# ── Scrivi log ────────────────────────────────────────────
def scrivi_log(bandi, nuovi_auto):
    OUTPUT_LOG.parent.mkdir(exist_ok=True)
    lines = [
        f"Ultimo aggiornamento: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"Bandi totali: {len(bandi)}",
        f"Verificati manualmente: {sum(1 for b in bandi if b.get('fonte')=='VERIFICATO')}",
        f"Trovati automaticamente: {nuovi_auto}",
        f"Aperti: {sum(1 for b in bandi if b.get('status')=='APERTO')}",
        f"In scadenza (<=30gg): {sum(1 for b in bandi if b.get('status')=='IN SCADENZA')}",
        "",
        "BANDI IN SCADENZA ENTRO 30 GIORNI:",
    ]
    for b in bandi:
        dl = days_left(b.get("scadenza", ""))
        if 0 <= dl <= 30:
            lines.append(f"  - {b['titolo']} (scade {b['scadenza']}, {dl}gg)")
    with open(OUTPUT_LOG, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ── Main ──────────────────────────────────────────────────
def main():
    print(f"\n{'='*55}")
    print(f"BandiRadar v2 — {TODAY.strftime('%d/%m/%Y')}")
    print(f"{'='*55}")

    # 1. Carica bandi verificati
    with open(DATA_FILE, encoding="utf-8") as f:
        bandi_base = json.load(f)
    print(f"\n[1] {len(bandi_base)} bandi verificati caricati da {DATA_FILE.name}")

    # 2. Scraping fonti esterne
    print("\n[2] Scraping fonti ufficiali...")
    scraped = []
    for source in SOURCES:
        print(f"  Scraping: {source['url']}")
        risultati = scrape_source(source)
        scraped.extend(risultati)
    print(f"  Totale trovati via scraping: {len(scraped)}")

    # 3. Merge e pulizia
    print("\n[3] Merge e calcolo status...")
    tutti = merge_bandi(bandi_base, scraped)
    nuovi_auto = sum(1 for b in tutti if b.get("fonte") != "VERIFICATO")
    print(f"  Totale finale: {len(tutti)} bandi ({nuovi_auto} automatici)")

    # 4. Genera output
    print("\n[4] Generazione output...")
    genera_excel(tutti, OUTPUT_XLSX)
    salva_json(tutti)
    scrivi_log(tutti, nuovi_auto)

    print(f"\nCompletato in {TODAY.strftime('%d/%m/%Y')}.")
    return 0

if __name__ == "__main__":
    sys.exit(main())
